import csv
import os
import shutil
import sys
import sqlite3


from PyQt5 import uic, QtCore
from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import QWidget, QApplication, QTableWidget, QTableWidgetItem, QComboBox, QListWidgetItem, \
    QMessageBox, QFileDialog, QLineEdit
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class SQL:
    def __init__(self):
        super(SQL, self).__init__()
        self.db = sqlite3.connect('database.db')
        self.sql_s("PRAGMA foreign_keys = OFF;")

    def sql_r(self, cmd):
        c = self.db.cursor()
        c.execute(cmd)
        return c.fetchall()

    def sql_s(self, cmd, data=None):
        c = self.db.cursor()
        if data:
            c.execute(cmd,data)
        else:
            c.execute(cmd)
        self.db.commit()

class ServicesWind(QWidget):
    def __init__(self, user=None):
        super(ServicesWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        uic.loadUi('services.ui',self)
        self.add_b.clicked.connect(self.add_serv)
        self.del_b.clicked.connect(self.delete_serv)
        self.services_table.cellPressed.connect(self.select_serv)
        self.back_b.clicked.connect(self.close)

        if self.auser[3] != 'Администратор':
            self.name_e.setEnabled(False)
            self.price_e.setEnabled(False)
            self.add_b.hide()
            self.del_b.hide()

        self.all_serv()

    def all_serv(self):
        table_name = 'Services'
        self.name_e.setText('')
        self.price_e.setText('')
        self.data = self.sql.sql_r(f"SELECT * FROM `{table_name}`")
        headers = self.sql.sql_r(f"PRAGMA table_info({table_name})")
        headers = [i[1] for i in headers]
        table = self.services_table
        #table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setRowCount(len(self.data))
        table.setHorizontalHeaderLabels(headers)

        for y, u in enumerate(self.data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))

    def select_serv(self,y,_):
        self.seluser = self.data[y]
        self.del_b.setEnabled(True)

    def add_serv(self):
        new_data = [self.name_e.text(), self.price_e.text(), self.count_cb.currentText()]

        self.sql.sql_s(f"INSERT INTO `Services` VALUES(null,?,?,?)", new_data)
        self.all_serv()

    def delete_serv(self):
        uid = self.seluser[0]
        self.sql.sql_s(f"DELETE FROM `Services` WHERE id={uid}")
        self.all_serv()





class MainWindow(QWidget):
    def __init__(self, user=None):
        super(MainWindow, self).__init__()
        self.auser = user
        print(self.auser)
        uic.loadUi('mgui.ui',self)
        self.dostup_l.setText('Доступ: ' + self.auser[3])
        self.users_b.clicked.connect(self.user_window)
        self.services_b.clicked.connect(self.services_window)
        self.pricelist_b.clicked.connect(self.pricelist_window)
        self.order_b.clicked.connect(self.order_window)
        self.check_b.clicked.connect(self.check_orders_window)
        self.extra_b.clicked.connect(self.extra_window)
        self.leave_b.clicked.connect(self.auth_window)

        if self.auser[3] == 'Менеджер':
            self.users_b.hide()
            self.services_b.hide()
            self.pricelist_b.hide()
            self.extra_b.hide()

    def auth_window(self):
        self.w = Auth()
        self.w.show()
        self.close()

    def user_window(self):
        self.w = UsersWind(self.auser)
        self.w.show()

    def services_window(self):
        self.w = ServicesWind(self.auser)
        self.w.show()

    def pricelist_window(self):
        self.w = PricelistWind(self.auser)
        self.w.show()

    def order_window(self):
        self.w = OrderWind(self.auser)
        self.w.show()

    def check_orders_window(self):
        self.w = CheckWind()
        self.w.show()

    def extra_window(self):
        self.w =ExtraWind()
        self.w.show()

class ExtraWind(QWidget):
    def __init__(self, user=None):
        super(ExtraWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        uic.loadUi('extra_menu.ui', self)
        self.export_db_b.clicked.connect(self.export_db)
        self.import_db_b.clicked.connect(self.import_db)
        self.export_csv_b.clicked.connect(self.export_csv)
        self.import_csv_b.clicked.connect(self.import_csv)
        self.back_b.clicked.connect(self.close)

    def export_db(self):
        if not os.path.exists('backup'):
            os.mkdir('backup')
        shutil.copy('database.db','backup/database_bk.db')

        self.mb = QMessageBox()
        self.mb.setText('Резервная копия сохранена в backup')
        self.mb.show()
        self.mb.exec()

    def import_db(self):
        if os.path.exists('backup/database_bk.db'):
            shutil.copy('backup/database_bk.db', './database.db')

            self.mb = QMessageBox()
            self.mb.setText('Резервная копия загружена из backup')
            self.mb.show()
            self.mb.exec()

    def export_csv(self):
        self.plus_w = QWidget()
        uic.loadUi('select_table.ui',self.plus_w)
        self.plus_w.act_b.clicked.connect(self.export_csv_conf)
        self.plus_w.act_b.setText('Экспортировать')
        self.plus_w.show()

        tables = self.sql.sql_r("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [t[0] for t in tables]
        tables.remove('sqlite_sequence')
        self.plus_w.table_cb.addItems(tables)

    def export_csv_conf(self):
        table = self.plus_w.table_cb.currentText()
        headers = self.sql.sql_r(f"PRAGMA table_info(`{table}`)")
        headers = [i[1] for i in headers]
        data = self.sql.sql_r(F"SELECT * FROM `{table}`")
        with open(f'backup/export_{table}.csv', 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)

        self.m = QMessageBox()
        self.m.setText(f'Экспорт CSV в {table} завершён!')
        self.m.exec()

        self.plus_w.close()

    def import_csv(self):
        self.plus_w = QWidget()
        uic.loadUi('select_table.ui',self.plus_w)
        self.plus_w.act_b.clicked.connect(self.import_csv_conf)
        self.plus_w.act_b.setText('Импортировать')
        self.plus_w.show()

        tables = self.sql.sql_r("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [t[0] for t in tables]
        tables.remove('sqlite_sequence')
        self.plus_w.table_cb.addItems(tables)

    def import_csv_conf(self):
        table = self.plus_w.table_cb.currentText()
        path, _ = QFileDialog.getOpenFileName(self,'Выберите файл','backup','CSV (*.csv)')
        if path:
            with open(path, 'r') as fin:
                dr = csv.DictReader(fin)
                student_info = [[v for _,v in i.items()] for i in dr]

            col_len = len(student_info[0])-1
            self.sql.sql_s(f"DELETE FROM `{table}`")
            for row in student_info:
                self.sql.sql_s(f"INSERT INTO `{table}` VALUES(?{',?'*col_len})", row)

            self.m = QMessageBox()
            self.m.setText(f'Импорт CSV в {table} завершён!')
            self.m.exec()

            self.plus_w.close()

class CheckWind(QWidget):
    def __init__(self, user=None):
        super(CheckWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        self.sortchk_sum = True
        self.sortchk_date = True
        self.filtered_data = []
        uic.loadUi('check_orders.ui',self)
        self.date1.dateChanged.connect(self.date_sort)
        self.date2.dateChanged.connect(self.date_sort)
        self.search_e.editingFinished.connect(self.search_sort)
        self.checkbox1.stateChanged.connect(self.status_sort)
        self.checkbox2.stateChanged.connect(self.status_sort)
        self.checkbox3.stateChanged.connect(self.status_sort)
        self.clearFil_b.clicked.connect(self.clear_filter)
        self.sortsum_b.clicked.connect(self.sort_sum)
        self.sortdate_b.clicked.connect(self.sort_date)
        self.sort_clear_b.clicked.connect(self.clear_sort)
        self.excel_b.clicked.connect(self.excel_export)
        self.info_b.clicked.connect(self.order_info)
        self.orders_table.cellClicked.connect(self.change_current_order)
        self.back_menu_b.clicked.connect(self.close)
        self.get_data()

    def get_data(self):
        self.table_name = 'Order'
        headers = self.sql.sql_r(f"PRAGMA table_info(`{self.table_name}`)")
        headers = [i[1] for i in headers]
        self.data = self.sql.sql_r(f"SELECT * FROM `{self.table_name}`")
        self.filtered_data = self.data
        table = self.orders_table
        #table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(self.data))
        for y, u in enumerate(self.data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))
        table.resizeColumnsToContents()
        table.setColumnWidth(3, 150)
        table.resizeRowsToContents()

        combobox = self.search_cb
        combobox.clear()
        combobox.addItems(headers)

    def upd_table(self, data):
        table = self.orders_table
        table.setRowCount(len(data))
        for y, u in enumerate(data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))
        table.resizeColumnsToContents()
        table.setColumnWidth(3, 150)
        table.resizeRowsToContents()

    def date_sort(self):
        table = self.orders_table
        datelist = [f for f in self.data if QDate.fromString(f[5], 'dd.MM.yyyy') >= self.date1.date()]
        datelist = [f for f in datelist if QDate.fromString(f[5], 'dd.MM.yyyy') <= self.date2.date()]

        self.upd_table(datelist)
        self.filtered_data = datelist

    def search_sort(self):
        col_id = self.search_cb.currentIndex()
        text =  self.search_e.text()
        searchlist = [s for s in self.data if text in str(s[col_id])]

        self.upd_table(searchlist)

    def status_sort(self):
        a, b, c = '#', '#', '#'
        if self.checkbox1.isChecked():
            a = self.checkbox1.text()
        if self.checkbox2.isChecked():
            b = self.checkbox2.text()
        if self.checkbox3.isChecked():
            c = self.checkbox3.text()
        if a=='#' and b=='#' and c=='#':
            self.date_sort()
            return 0
        l = [a,b,c]
        self.date_sort()
        searchlist = [s for s in self.filtered_data if any(x in l for x in s)]

        self.upd_table(searchlist)
        self.filtered_data = searchlist

    def clear_filter(self):
        self.checkbox1.setChecked(False)
        self.checkbox1.setChecked(False)
        self.checkbox1.setChecked(False)
        self.date1.setDate(QDate(2000,1,1))
        self.date2.setDate(QDate(2099,1,1))


    def sort_sum(self):
        sorteddata = sorted(self.filtered_data, key=lambda x: x[7], reverse=self.sortchk_sum)
        if self.sortchk_sum:
            self.sortsum_b.setText('Сумма заказа ↓')
            self.sortchk_sum = False
        else:
            self.sortsum_b.setText('Сумма заказа ↑')
            self.sortchk_sum = True

        self.upd_table(sorteddata)

    def sort_date(self):
        sorteddata = sorted(self.filtered_data, key=lambda x: x[5], reverse=self.sortchk_date)
        if self.sortchk_date:
            self.sortdate_b.setText('Дата выдачи ↓')
            self.sortchk_date = False
        else:
            self.sortdate_b.setText('Дата выдачи ↑')
            self.sortchk_date = True

        self.upd_table(sorteddata)

    def clear_sort(self):
        self.upd_table(self.filtered_data)

    def excel_export(self):
        workbook = Workbook()
        sheet = workbook.active

        headers = self.sql.sql_r(f"PRAGMA table_info(`{self.table_name}`)")
        headers = [i[1] for i in headers]
        sheet.append(headers)

        for row in self.data:
            sheet.append(row)

        column_widths = [4, 29, 13, 30, 12, 12, 8, 12]
        for i, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width
            sheet[get_column_letter(i)+'1'].font = Font(bold=True)

        workbook.save("Отчёт.xlsx")

        self.m = QMessageBox()
        self.m.setText('Отчёт сохранён в каталоге!')
        self.m.setWindowTitle('Отчёт')
        self.m.exec()

    def change_current_order(self, y, _):
        self.selected_order = self.data[y]
        self.info_b.setEnabled(True)

    def order_info(self):
        self.w = QWidget()
        uic.loadUi('order_info.ui', self.w)
        self.w.show()
        self.w.status_cb.currentIndexChanged.connect(self.change_status)
        self.w.back_b.clicked.connect(self.w.close)
        self.w.price_l.setText('Общая стоимость: ' + str(self.selected_order[7]))

        table_name = 'ServicesInOrder'
        headers = self.sql.sql_r(f"PRAGMA table_info({table_name})")
        labels = [i[1] for i in headers][2:]
        table = self.w.carpet_table
        #table = QTableWidget()
        table.setColumnCount(len(labels))
        table.setHorizontalHeaderLabels(labels)

        carpet_data = self.sql.sql_r(f"SELECT * FROM {table_name} WHERE ord_id={self.selected_order[0]}")
        carpet_data = [i[2:] for i in carpet_data]

        table.setRowCount(len(carpet_data))
        for y, u in enumerate(carpet_data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))
        table.resizeRowsToContents()


    def change_status(self):
        stat = self.w.status_cb.currentText()
        ord_id = self.selected_order[0]
        self.sql.sql_s(f"UPDATE `{self.table_name}` SET Статус='{stat}' WHERE id={ord_id}")
        self.data = self.sql.sql_r(f"SELECT * FROM `{self.table_name}`")
        self.filtered_data = self.data
        self.upd_table(self.data)



class OrderWind(QWidget):
    def __init__(self, user=None):
        super(OrderWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        uic.loadUi('order.ui',self)
        self.add_b.clicked.connect(self.add_serv)
        self.del_b.clicked.connect(self.delete_serv)
        self.carpet_table.cellPressed.connect(self.select_serv)
        self.w_sb.valueChanged.connect(self.count_s)
        self.h_sb.valueChanged.connect(self.count_s)
        self.carpet_cb.currentIndexChanged.connect(self.current_price)
        self.s_l.textChanged.connect(self.result_price)
        self.order_b.clicked.connect(self.change_wind)
        self.create_ord_b.clicked.connect(self.create_order)
        self.back_b.clicked.connect(self.back)
        self.back_b_2.clicked.connect(self.close)
        self.carpets_in_order = []
        self.get_data()

    def get_data(self):
        self.table_name = 'ServicesInOrder'
        headers = self.sql.sql_r(f"PRAGMA table_info({self.table_name})")
        headers = [i[1] for i in headers][2:]
        table = self.carpet_table
        #table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)


        self.pricelist = self.sql.sql_r('SELECT * FROM `Pricelist`')
        self.servlist = self.sql.sql_r('SELECT * FROM `Services`')
        p_names = [n[1] for n in self.pricelist]
        self.carpet_cb.addItems(p_names)

        s_names = [n[1] for n in self.servlist]
        self.carpet_list.clear()
        for s in s_names:
            item = QListWidgetItem(s)
            item.setFlags(item.flags()| QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.carpet_list.addItem(item)

    def current_price(self,ind):
        self.current_carpet = self.pricelist[ind]
        self.pricem_l.setText(str(self.pricelist[ind][2]))

    def count_s(self):
        w = self.w_sb.value()
        h = self.h_sb.value()
        res = w * h
        self.s_l.setText(str(res))

    def result_price(self):
        if self.pricem_l.text() and self.s_l.text():
            result = float(self.pricem_l.text()) * float(self.s_l.text())
            self.resprice_l.setText(str(result))

    def add_serv(self):
        c = self.carpet_list.count()
        lil = [self.carpet_list.item(x) for x in range(c)]
        serv = []
        ser_price = 0
        for i in lil:
            if i.checkState():
                serv.append(i.text())
                for s in self.servlist:
                    if i.text() in s:
                        if s[3] == '%':
                            ser_price += float(self.resprice_l.text()) * (s[2]/100)
                        else:
                            ser_price += s[2]


        carpet = self.carpet_cb.currentText()
        w = self.w_sb.value()
        h = self.h_sb.value()
        cl_price = self.resprice_l.text()
        final_price = float(cl_price) + ser_price
        serv_str = ''
        for s in serv:
            serv_str += ' '+s
        self.carpets_in_order.append([carpet,w,h,cl_price,ser_price,final_price,serv_str])
        self.upd_table()
        self.order_b.setEnabled(True)

    def upd_table(self):
        self.carpet_table.setRowCount(len(self.carpets_in_order))
        for y, u in enumerate(self.carpets_in_order):
            for x, e in enumerate(u):
                self.carpet_table.setItem(y,x,QTableWidgetItem(str(e)))
        self.carpet_table.resizeRowsToContents()

    def select_serv(self,y):
        self.seluser = y
        self.del_b.setEnabled(True)

    def delete_serv(self):
        id = self.seluser
        self.carpet_table.removeRow(id)
        self.carpets_in_order.pop(id)
        self.upd_table()
        if not self.carpets_in_order:
            self.del_b.setEnabled(False)

    def change_wind(self):
        self.table_name = 'ServicesInOrder'
        headers = self.sql.sql_r(f"PRAGMA table_info({self.table_name})")
        headers = [i[1] for i in headers][2:]
        table = self.carpet_table_2
        #table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        self.stackedWidget.setCurrentWidget(self.page_2)
        self.carpet_table_2.setRowCount(len(self.carpets_in_order))
        for y, u in enumerate(self.carpets_in_order):
            for x, e in enumerate(u):
                self.carpet_table_2.setItem(y, x, QTableWidgetItem(str(e)))
        self.carpet_table_2.resizeRowsToContents()
        all_price = [p[5] for p in self.carpets_in_order]
        self.final_price = sum(all_price)
        self.price_l.setText('Общая стоимость: ' + str(self.final_price))
        self.dateEdit.setMinimumDate(QtCore.QDate.currentDate())

    def back(self):
        self.stackedWidget.setCurrentWidget(self.page)

    def create_order(self):
        name = self.name_l.text()
        phone = self.phone_l.text()
        addres = self.address_l.text()
        date_a = QtCore.QDate.currentDate().toString('dd.MM.yyyy')
        date_b = self.dateEdit.text()
        this_order = [name, phone, addres, date_a, date_b, 'Принят', str(self.final_price), str(self.auser[0])]
        print(this_order)
        l = len(this_order)
        self.sql.sql_s(f"INSERT INTO `Order` VALUES(null{l * ',?'})", this_order)
        ord_id = self.sql.sql_r("select max(id) from `Order`")[0][0]
        c_l = len(self.carpets_in_order[0])
        for carpet in self.carpets_in_order:
            self.sql.sql_s(f"INSERT INTO `{self.table_name}` VALUES(null,{ord_id}{',?'*c_l})", carpet)

        self.mb = QMessageBox()
        self.mb.setText('Заказ успешно создан')
        self.mb.show()
        self.mb.exec()

        self.reset_wind()
        self.back()

    def reset_wind(self):
        self.resprice_l.setText('')
        self.w_sb.setValue(0)
        self.h_sb.setValue(0)
        self.s_l.setText('')
        self.carpet_table.clear()
        self.carpet_cb.clear()
        self.del_b.setEnabled(False)
        self.order_b.setEnabled(False)
        self.carpets_in_order = []
        self.name_l.setText('')
        self.phone_l.setText('')
        self.address_l.setText('')
        self.get_data()


class PricelistWind(QWidget):
    def __init__(self, user=None):
        super(PricelistWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        uic.loadUi('pricelist.ui',self)
        self.add_b.clicked.connect(self.add_serv)
        self.del_b.clicked.connect(self.delete_serv)
        self.services_table.cellPressed.connect(self.select_serv)
        self.back_b.clicked.connect(self.close)

        if self.auser[3] != 'Администратор':
            self.name_e.setEnabled(False)
            self.price_e.setEnabled(False)
            self.add_b.hide()
            self.del_b.hide()

        self.all_carpet()

    def all_carpet(self):
        self.table_name = 'Pricelist'
        self.name_e.setText('')
        self.price_e.setText('')
        self.data = self.sql.sql_r(f"SELECT * FROM `{self.table_name}`")
        headers = self.sql.sql_r(f"PRAGMA table_info({self.table_name})")
        headers = [i[1] for i in headers]
        table = self.services_table
        #table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setRowCount(len(self.data))
        table.setHorizontalHeaderLabels(headers)

        for y, u in enumerate(self.data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))

    def select_serv(self,y,_):
        self.seluser = self.data[y]
        self.del_b.setEnabled(True)

    def add_serv(self):
        new_data = [self.name_e.text(), self.price_e.text()]
        l = len(new_data)
        self.sql.sql_s(f"INSERT INTO `{self.table_name}` VALUES(null{l*',?'})", new_data)
        self.all_carpet()

    def delete_serv(self):
        uid = self.seluser[0]
        self.sql.sql_s(f"DELETE FROM `{self.table_name}` WHERE id={uid}")
        self.all_carpet()



class UsersWind(QWidget):
    def __init__(self, user=None):
        super(UsersWind, self).__init__()
        self.auser = user
        self.sql = SQL()
        uic.loadUi('users.ui',self)
        self.add_b.clicked.connect(self.add_user)
        self.del_b.clicked.connect(self.delete_user)
        self.user_table.cellPressed.connect(self.select_user)
        self.back_b.clicked.connect(self.close)

        if self.auser[3] != 'Администратор':
            self.login_e.setEnabled(False)
            self.pass_e.setEnabled(False)
            self.add_b.hide()
            self.del_b.hide()

        self.all_users()

    def all_users(self):
        table_name = 'Users'
        self.login_e.setText('')
        self.pass_e.setText('')
        self.data = self.sql.sql_r(f"SELECT * FROM `{table_name}`")
        headers = self.sql.sql_r(f"PRAGMA table_info({table_name})")
        headers = [i[1] for i in headers]
        table = self.user_table
        #table = QTableWidget()
        table.setColumnCount(len(self.data[0]))
        table.setRowCount(len(self.data))
        table.setHorizontalHeaderLabels(headers)

        for y, u in enumerate(self.data):
            for x, e in enumerate(u):
                table.setItem(y,x,QTableWidgetItem(str(e)))

        lvllist = self.sql.sql_r(f"SELECT * FROM `Levels`")
        lvllist = [i[1] for i in lvllist]
        #combobox = QComboBox(self)
        combobox = self.lvl_cb
        combobox.clear()
        combobox.addItems(lvllist)

    def select_user(self,y,_):
        self.seluser = self.data[y]
        self.del_b.setEnabled(True)

    def add_user(self):
        new_data = [ self.login_e.text(), self.pass_e.text(), self.lvl_cb.currentText()]
        ids = [i[0] for i in self.data]
        if not ids:
            return 0
        curid = max(ids) + 1
        self.sql.sql_s(f"INSERT INTO Users VALUES({curid},?,?,?)", new_data)
        self.all_users()

    def delete_user(self):
        uid = self.seluser[0]
        self.sql.sql_s(f"DELETE FROM `Users` WHERE id={uid}")
        self.all_users()

class Auth(QWidget):
    def __init__(self):
        super(Auth, self).__init__()
        uic.loadUi('auth.ui',self)
        self.sql = SQL()
        self.login_b.clicked.connect(self.try_enter)
        self.show_pass_cb.clicked.connect(self.show_pass)

    def show_pass(self):
        if self.show_pass_cb.isChecked():
            self.pass_e.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            self.pass_e.setEchoMode(QLineEdit.EchoMode.Password)

    def try_enter(self):
        login = self.login_e.text()
        password = self.pass_e.text()

        all_users = self.sql.sql_r("SELECT * FROM Users")
        for u in all_users:
            if login == u[1] and password == u[2]:
                self.w = MainWindow(u)
                self.w.show()
                self.close()
                return 0

        self.m = QMessageBox()
        self.m.setText('Не верный пароль или логин!')
        self.m.exec()


if __name__ == '__main__':
    apka = QApplication(sys.argv)
    w = Auth()
    w.show()
    apka.exec()
